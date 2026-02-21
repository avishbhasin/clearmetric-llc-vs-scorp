"""
ClearMetric LLC vs S-Corp Tax Calculator — Premium Excel Template
Product for Gumroad ($16.99)

3 Sheets:
  1. LLC vs S-Corp Comparison — inputs + both calculations side by side
  2. Break-Even Analysis — income $50K to $300K in $10K steps
  3. How To Use — instructions

Design: Indigo palette (#2C3E8F primary, #1A2766 dark, #D6DBEF input)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# DESIGN SYSTEM — Indigo
# ============================================================
PRIMARY = "2C3E8F"
DARK = "1A2766"
WHITE = "FFFFFF"
INPUT_COLOR = "D6DBEF"
LIGHT_GRAY = "F5F6FA"
MED_GRAY = "D5D8DC"
DARK_GRAY = "5D6D7E"
LIGHT_INDIGO = "E8EAF6"
ACCENT = "3F51B5"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="B0BEC5", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color="2C3E50")
FONT_INPUT = Font(name="Calibri", size=12, color=DARK, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color="2C3E50")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=DARK)
FONT_SMALL = Font(name="Calibri", size=9, color=DARK_GRAY, italic=True)

FILL_PRIMARY = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_INDIGO, end_color=LIGHT_INDIGO, fill_type="solid")

THIN = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_PRIMARY
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_PRIMARY
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def label_calc(ws, row, lc, vc, label, formula, fmt=None, bold=False):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=formula)
    cv.font = FONT_BOLD if bold else FONT_VALUE
    cv.fill = FILL_WHITE
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# SHEET 1: LLC vs S-Corp Comparison
# ============================================================
def build_comparison(ws):
    ws.title = "LLC vs S-Corp"
    ws.sheet_properties.tabColor = PRIMARY
    cols(ws, {"A": 2, "B": 28, "C": 14, "D": 4, "E": 28, "F": 14, "G": 2})

    for r in range(1, 55):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:F1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:F2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="LLC vs S-CORP TAX CALCULATOR")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(
        row=3, column=2,
        value="Enter your numbers in the indigo cells. Both paths calculate side by side.",
    )
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # ===== INPUTS =====
    header_bar(ws, 5, 2, 3, "INPUTS")
    label_input(ws, 6, 2, 3, "Business Net Income ($)", 120000, "$#,##0")
    label_input(ws, 7, 2, 3, "Other W-2 Income ($)", 0, "$#,##0")
    label_input(ws, 8, 2, 3, "Filing: 1=Single 2=MFJ 3=HOH", 1, "0")
    label_input(ws, 9, 2, 3, "S-Corp Reasonable Salary ($)", 60000, "$#,##0")
    label_input(ws, 10, 2, 3, "State Tax Rate (e.g. 0.093)", 0.093, "0.0%")
    label_input(ws, 11, 2, 3, "Health Insurance ($)", 6000, "$#,##0")
    label_input(ws, 12, 2, 3, "Retirement Contribution ($)", 20000, "$#,##0")
    label_input(ws, 13, 2, 3, "QBI Eligible? (1=Yes)", 1, "0")

    # Standard deduction helper
    ws.cell(row=14, column=2, value="Std Ded (auto)").font = FONT_LABEL
    ws.cell(row=14, column=2).fill = FILL_GRAY
    ws.cell(row=14, column=2).border = THIN
    ws.cell(row=14, column=3, value="=IF(C8=1,16100,IF(C8=2,32200,24150))")
    ws.cell(row=14, column=3).font = FONT_VALUE
    ws.cell(row=14, column=3).fill = FILL_WHITE
    ws.cell(row=14, column=3).number_format = "$#,##0"
    ws.cell(row=14, column=3).border = THIN

    # ===== LLC COLUMN (E) =====
    header_bar(ws, 5, 5, 6, "LLC (Sole Prop)", FILL_DARK)
    label_calc(ws, 6, 5, 6, "SE Taxable (92.35%)", "=C6*0.9235", "$#,##0")
    label_calc(ws, 7, 5, 6, "SE Tax (15.3%)", "=F6*0.153", "$#,##0")
    label_calc(ws, 8, 5, 6, "SE Deduction (50%)", "=F7*0.5", "$#,##0")
    label_calc(ws, 9, 5, 6, "AGI", "=C7+C6-C11-C12-F8", "$#,##0")
    label_calc(ws, 10, 5, 6, "Taxable (before QBI)", "=MAX(0,F9-C14)", "$#,##0")
    label_calc(ws, 11, 5, 6, "QBI Ded (20% if elig)", "=IF(C13=1,MIN(C6*0.2,F10*0.2),0)", "$#,##0")
    label_calc(ws, 12, 5, 6, "Taxable Income", "=MAX(0,F10-F11)", "$#,##0", bold=True)
    label_calc(ws, 13, 5, 6, "Federal Tax (est)", "=F12*0.22", "$#,##0")
    label_calc(ws, 14, 5, 6, "State Tax", "=F12*C10", "$#,##0")
    label_calc(ws, 15, 5, 6, "LLC Total Tax", "=F7+F13+F14", "$#,##0", bold=True)

    # ===== S-CORP COLUMN (E+F for S-Corp, use cols 5-6 merged or separate)
    # We need another set. Use columns 5-6 for LLC. Add S-Corp in 7? No, we have 7 cols. Use E and F.
    # Actually we have B,C for inputs. D is spacer. E,F for LLC. We need G for S-Corp? Let me use E for LLC labels, F for LLC values. Then we need another pair. Use 5-6 for both - actually we have 6 columns (B through G). So: B,C = inputs. E,F = LLC. We need S-Corp. Let me add columns - we have A(2), B(28), C(14), D(4), E(28), F(14), G(2). So E,F are used. Add H,I for S-Corp.
    cols(ws, {"A": 2, "B": 24, "C": 12, "D": 3, "E": 24, "F": 12, "G": 3, "H": 24, "I": 12, "J": 2})

    header_bar(ws, 5, 8, 9, "S-CORP", FILL_DARK)
    # S-Corp: salary, employer FICA, distributions
    ws.cell(row=6, column=8, value="Salary (capped)").font = FONT_LABEL
    ws.cell(row=6, column=8).fill = FILL_GRAY
    ws.cell(row=6, column=8).border = THIN
    ws.cell(row=6, column=9, value="=MIN(C9,C6/1.0765)").font = FONT_VALUE
    ws.cell(row=6, column=9).fill = FILL_WHITE
    ws.cell(row=6, column=9).number_format = "$#,##0"
    ws.cell(row=6, column=9).border = THIN
    ws.cell(row=6, column=9).alignment = ALIGN_R

    ws.cell(row=7, column=8, value="Employer FICA (7.65%)").font = FONT_LABEL
    ws.cell(row=7, column=8).fill = FILL_GRAY
    ws.cell(row=7, column=8).border = THIN
    ws.cell(row=7, column=9, value="=I6*0.0765").font = FONT_VALUE
    ws.cell(row=7, column=9).fill = FILL_WHITE
    ws.cell(row=7, column=9).number_format = "$#,##0"
    ws.cell(row=7, column=9).border = THIN
    ws.cell(row=7, column=9).alignment = ALIGN_R

    ws.cell(row=8, column=8, value="Distributions").font = FONT_LABEL
    ws.cell(row=8, column=8).fill = FILL_GRAY
    ws.cell(row=8, column=8).border = THIN
    ws.cell(row=8, column=9, value="=MAX(0,C6-I6-I7)").font = FONT_VALUE
    ws.cell(row=8, column=9).fill = FILL_WHITE
    ws.cell(row=8, column=9).number_format = "$#,##0"
    ws.cell(row=8, column=9).border = THIN
    ws.cell(row=8, column=9).alignment = ALIGN_R

    ws.cell(row=9, column=8, value="Total FICA (15.3%)").font = FONT_LABEL
    ws.cell(row=9, column=8).fill = FILL_GRAY
    ws.cell(row=9, column=8).border = THIN
    ws.cell(row=9, column=9, value="=I6*0.153").font = FONT_VALUE
    ws.cell(row=9, column=9).fill = FILL_WHITE
    ws.cell(row=9, column=9).number_format = "$#,##0"
    ws.cell(row=9, column=9).border = THIN
    ws.cell(row=9, column=9).alignment = ALIGN_R

    ws.cell(row=10, column=8, value="AGI").font = FONT_LABEL
    ws.cell(row=10, column=8).fill = FILL_GRAY
    ws.cell(row=10, column=8).border = THIN
    ws.cell(row=10, column=9, value="=C7+I6+I8-C11-C12").font = FONT_VALUE
    ws.cell(row=10, column=9).fill = FILL_WHITE
    ws.cell(row=10, column=9).number_format = "$#,##0"
    ws.cell(row=10, column=9).border = THIN
    ws.cell(row=10, column=9).alignment = ALIGN_R

    ws.cell(row=11, column=8, value="Taxable (before QBI)").font = FONT_LABEL
    ws.cell(row=11, column=8).fill = FILL_GRAY
    ws.cell(row=11, column=8).border = THIN
    ws.cell(row=11, column=9, value="=MAX(0,I10-C14)").font = FONT_VALUE
    ws.cell(row=11, column=9).fill = FILL_WHITE
    ws.cell(row=11, column=9).number_format = "$#,##0"
    ws.cell(row=11, column=9).border = THIN
    ws.cell(row=11, column=9).alignment = ALIGN_R

    ws.cell(row=12, column=8, value="QBI Ded (20% of dist)").font = FONT_LABEL
    ws.cell(row=12, column=8).fill = FILL_GRAY
    ws.cell(row=12, column=8).border = THIN
    ws.cell(row=12, column=9, value="=IF(C13=1,MIN(I8*0.2,I11*0.2),0)").font = FONT_VALUE
    ws.cell(row=12, column=9).fill = FILL_WHITE
    ws.cell(row=12, column=9).number_format = "$#,##0"
    ws.cell(row=12, column=9).border = THIN
    ws.cell(row=12, column=9).alignment = ALIGN_R

    ws.cell(row=13, column=8, value="Taxable Income").font = FONT_LABEL
    ws.cell(row=13, column=8).fill = FILL_GRAY
    ws.cell(row=13, column=8).border = THIN
    ws.cell(row=13, column=9, value="=MAX(0,I11-I12)").font = FONT_BOLD
    ws.cell(row=13, column=9).fill = FILL_WHITE
    ws.cell(row=13, column=9).number_format = "$#,##0"
    ws.cell(row=13, column=9).border = THIN
    ws.cell(row=13, column=9).alignment = ALIGN_R

    ws.cell(row=14, column=8, value="Federal Tax (est)").font = FONT_LABEL
    ws.cell(row=14, column=8).fill = FILL_GRAY
    ws.cell(row=14, column=8).border = THIN
    ws.cell(row=14, column=9, value="=I13*0.22").font = FONT_VALUE
    ws.cell(row=14, column=9).fill = FILL_WHITE
    ws.cell(row=14, column=9).number_format = "$#,##0"
    ws.cell(row=14, column=9).border = THIN
    ws.cell(row=14, column=9).alignment = ALIGN_R

    ws.cell(row=15, column=8, value="State Tax").font = FONT_LABEL
    ws.cell(row=15, column=8).fill = FILL_GRAY
    ws.cell(row=15, column=8).border = THIN
    ws.cell(row=15, column=9, value="=I13*C10").font = FONT_VALUE
    ws.cell(row=15, column=9).fill = FILL_WHITE
    ws.cell(row=15, column=9).number_format = "$#,##0"
    ws.cell(row=15, column=9).border = THIN
    ws.cell(row=15, column=9).alignment = ALIGN_R

    ws.cell(row=16, column=8, value="S-Corp Tax").font = FONT_LABEL
    ws.cell(row=16, column=8).fill = FILL_GRAY
    ws.cell(row=16, column=8).border = THIN
    ws.cell(row=16, column=9, value="=I9+I14+I15").font = FONT_VALUE
    ws.cell(row=16, column=9).fill = FILL_WHITE
    ws.cell(row=16, column=9).number_format = "$#,##0"
    ws.cell(row=16, column=9).border = THIN
    ws.cell(row=16, column=9).alignment = ALIGN_R

    ws.cell(row=17, column=8, value="S-Corp Costs").font = FONT_LABEL
    ws.cell(row=17, column=8).fill = FILL_GRAY
    ws.cell(row=17, column=8).border = THIN
    ws.cell(row=17, column=9, value=3500).font = FONT_VALUE
    ws.cell(row=17, column=9).fill = FILL_WHITE
    ws.cell(row=17, column=9).number_format = "$#,##0"
    ws.cell(row=17, column=9).border = THIN
    ws.cell(row=17, column=9).alignment = ALIGN_R

    ws.cell(row=18, column=8, value="S-Corp Total Cost").font = FONT_HEADER
    ws.cell(row=18, column=8).fill = FILL_DARK
    ws.cell(row=18, column=8).border = THIN
    ws.cell(row=18, column=9, value="=I16+I17").font = FONT_BOLD
    ws.cell(row=18, column=9).fill = FILL_LIGHT
    ws.cell(row=18, column=9).number_format = "$#,##0"
    ws.cell(row=18, column=9).border = THIN
    ws.cell(row=18, column=9).alignment = ALIGN_R

    # Verdict row
    ws.merge_cells("B20:I20")
    verdict = ws.cell(row=20, column=2, value="=IF(F15>I18,\"S-Corp saves $\"&TEXT(F15-I18,\"#,##0\"),\"LLC saves $\"&TEXT(I18-F15,\"#,##0\"))")
    verdict.font = Font(name="Calibri", size=14, bold=True, color=DARK)
    verdict.fill = FILL_LIGHT
    verdict.border = THIN
    verdict.alignment = ALIGN_C

    ws.protection.sheet = True
    input_cells = [(6, 3), (7, 3), (8, 3), (9, 3), (10, 3), (11, 3), (12, 3), (13, 3)]
    for r, c in input_cells:
        ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: Break-Even Analysis
# ============================================================
def build_breakeven(wb):
    ws = wb.create_sheet("Break-Even Analysis")
    ws.sheet_properties.tabColor = ACCENT
    cols(ws, {"A": 2, "B": 18, "C": 18, "D": 18, "E": 18})

    for r in range(1, 35):
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    ws.merge_cells("B1:D2")
    ws.cell(row=1, column=2, value="BREAK-EVEN ANALYSIS").font = FONT_TITLE
    ws.cell(row=1, column=2).fill = FILL_DARK
    ws.cell(row=1, column=2).alignment = ALIGN_C
    for r in range(1, 3):
        for c in range(2, 5):
            ws.cell(row=r, column=c).fill = FILL_DARK

    ws.merge_cells("B3:D3")
    ws.cell(row=3, column=2, value="Income from $50K to $300K in $10K steps. Uses inputs from Sheet 1.").font = FONT_SMALL
    ws.cell(row=3, column=2).alignment = ALIGN_L

    # Headers
    for col, label in [(2, "Business Income"), (3, "LLC Tax"), (4, "S-Corp Cost"), (5, "Savings (S-Corp)")]:
        c = ws.cell(row=5, column=col, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_PRIMARY
        c.border = THIN
        c.alignment = ALIGN_C

    # Data rows - simplified formulas referencing 'LLC vs S-Corp' sheet
    # Income in col B: 50000, 60000, ..., 300000
    sh = "'LLC vs S-Corp'!"
    for i in range(26):
        row = 6 + i
        inc = 50000 + i * 10000
        ws.cell(row=row, column=2, value=inc)
        ws.cell(row=row, column=2).number_format = "$#,##0"
        ws.cell(row=row, column=2).font = FONT_VALUE
        ws.cell(row=row, column=2).border = THIN
        # Tax base (income + w2 - health - retirement - std ded)
        tax_base = f"MAX(0,B{row}+{sh}C7-{sh}C11-{sh}C12-{sh}C14)"
        # LLC: SE tax + federal + state (simplified)
        ws.cell(row=row, column=3, value=f"=B{row}*0.9235*0.153+{tax_base}*0.22+{tax_base}*{sh}C10")
        ws.cell(row=row, column=3).number_format = "$#,##0"
        ws.cell(row=row, column=3).font = FONT_VALUE
        ws.cell(row=row, column=3).border = THIN
        # S-Corp: FICA on salary + same income tax + costs
        ws.cell(row=row, column=4, value=f"=MIN(B{row}*0.5,B{row}/1.0765)*0.153+{tax_base}*0.22+{tax_base}*{sh}C10+3500")
        ws.cell(row=row, column=4).number_format = "$#,##0"
        ws.cell(row=row, column=4).font = FONT_VALUE
        ws.cell(row=row, column=4).border = THIN
        # Savings (LLC - S-Corp, positive = S-Corp saves)
        ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        ws.cell(row=row, column=5).number_format = "$#,##0"
        ws.cell(row=row, column=5).font = FONT_VALUE
        ws.cell(row=row, column=5).border = THIN

    ws.protection.sheet = True
    for r in range(6, 32):
        ws.cell(row=r, column=2).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 3: How To Use
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = DARK_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE LLC vs S-CORP CALCULATOR")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'LLC vs S-Corp' tab and enter your numbers in the INDIGO (light blue) cells",
            "2. Business net income, other W-2 income, filing status, S-Corp reasonable salary",
            "3. State tax rate (e.g., 0.093 for California 9.3%), health insurance, retirement",
            "4. QBI eligible = 1 if you qualify for the 20% pass-through deduction",
            "5. LLC and S-Corp columns update automatically",
            "6. Use 'Break-Even Analysis' to see when S-Corp saves at different income levels",
        ]),
        ("INPUT EXPLANATIONS", [
            "Business Net Income: Profit after expenses (revenue minus deductible costs)",
            "S-Corp Reasonable Salary: Must be 'reasonable' per IRS — typically 30–50% of profit",
            "Health Insurance: Self-employed health insurance deduction (above-the-line)",
            "Retirement: SEP-IRA (max 25% of net) or Solo 401k contribution",
            "QBI: Qualified Business Income — 20% deduction for pass-through businesses",
            "State Tax Rate: Your state's income tax rate. No-tax states = 0",
        ]),
        ("LLC (SOLE PROP) PATH", [
            "All net income subject to self-employment tax (15.3% on 92.35%)",
            "SE tax includes Social Security (capped) + Medicare",
            "50% of SE tax reduces AGI",
            "QBI deduction: 20% of qualified income (if eligible)",
        ]),
        ("S-CORP PATH", [
            "Only salary subject to FICA (7.65% employee + 7.65% employer)",
            "Distributions (profit minus salary minus employer FICA) not subject to SE tax",
            "QBI deduction applies to distributions",
            "Additional costs: payroll ~$1,200, tax prep ~$1,500, state filing ~$800",
        ]),
        ("BREAK-EVEN", [
            "S-Corp typically becomes worthwhile when business net income exceeds ~$60K–$80K",
            "Below that, LLC costs less because S-Corp fees outweigh SE tax savings",
            "Your exact break-even depends on salary, state, and other income",
        ]),
        ("IMPORTANT NOTES", [
            "This is an estimator only — consult a CPA for your specific situation",
            "Federal tax uses simplified effective rate; actual brackets may differ",
            "SS wage base 2026: $184,500",
            "© 2026 ClearMetric. For educational use only. Not financial or tax advice.",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=DARK)
        ws.cell(row=r, column=2).fill = FILL_LIGHT
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color="2C3E50")
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building LLC vs S-Corp sheet...")
    build_comparison(ws)

    print("Building Break-Even Analysis sheet...")
    build_breakeven(wb)

    print("Building How To Use sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "output",
        "ClearMetric-LLC-vs-SCorp-Calculator.xlsx",
    )
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
